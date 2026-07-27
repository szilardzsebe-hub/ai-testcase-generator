from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(test_cases, file_path):
    """
    Export generated test cases to a professionally formatted Excel workbook.

    The exported workbook is intended for QA Engineers, Test Managers
    and other stakeholders while also serving as a foundation for
    future enterprise integrations.
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Cases"

    # Create descriptive column headers matching the internal
    # enterprise test case model.
    sheet.append([
        "TC ID",
        "Title",
        "Requirement",
        "Category",
        "Test Type",
        "Priority",
        "Preconditions",
        "Test Data",
        "Steps",
        "Expected Result",
        "Postconditions",
        "Automation Candidate",
        "Tags",
    ])

    # Apply professional formatting to the header row.
    header_font = Font(bold=True, color="FFFFFF")

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="4F81BD"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Export every generated test case.
    for tc in test_cases:

        sheet.append([
            tc.tc_id,
            tc.title,
            tc.requirement,
            tc.category.value,
            tc.tc_type.value,
            tc.priority.value,
            "\n".join(tc.preconditions),

            "\n".join(
                f"{key}: {value}"
                for key, value in tc.test_data.items()
            ),

            "\n".join(
                f"{index + 1}. {step}"
                for index, step in enumerate(tc.steps)
            ),

            tc.expected_result,

            "\n".join(tc.postconditions),

            tc.automation_candidate.value,

            ", ".join(tc.tags),
        ])

    # Enable text wrapping because many fields contain
    # multi-line information.
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

    # Keep the header visible while scrolling.
    sheet.freeze_panes = "A2"

    # Enable filtering for all columns.
    sheet.auto_filter.ref = sheet.dimensions

    # Automatically resize columns while limiting
    # excessive column widths.
    for column in sheet.columns:

        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        sheet.column_dimensions[column_letter].width = min(
            max_length + 2,
            40
        )

    # Increase row height for better readability.
    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 60

    # Apply borders to all populated cells.
    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border

    workbook.save(file_path)